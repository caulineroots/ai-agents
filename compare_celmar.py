import openpyxl

path1 = r'C:\Users\AVELL\Downloads\Cauline Roots\Celmar\254_BLN_Planilha Civil.xlsx'
path2 = r'C:\Users\AVELL\Downloads\Cauline Roots\Celmar\1ª Proposta CELMAR BLN.xlsx'

wb1 = openpyxl.load_workbook(path1, data_only=True)
ws1 = wb1['Planilha2']

wb2 = openpyxl.load_workbook(path2, data_only=True)
ws2 = wb2['1ª Proposta']


def parse_sheet(ws, start_row=30):
    items = {}
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        item_code = row[2]
        if not isinstance(item_code, str) or '.' not in item_code:
            continue
        desc  = str(row[3] or '').strip()
        unit  = str(row[4] or '').strip()
        qde   = row[5]
        mat   = row[6]
        mo    = row[7]
        total = row[8]
        code = item_code.strip()
        if code not in items:  # primeiro encontrado vence (evita duplicatas)
            items[code] = {
                'desc': desc, 'unit': unit,
                'qde': qde if qde is not None else 0,
                'mat': mat if mat is not None else 0,
                'mo':  mo  if mo  is not None else 0,
                'total': total if total is not None else 0,
            }
    return items


def sort_key(code):
    parts = []
    for p in code.split('.'):
        try:
            parts.append((0, int(p)))
        except ValueError:
            parts.append((1, p))
    return parts


ref  = parse_sheet(ws1)   # Planilha Civil (gabarito com QDEs)
prop = parse_sheet(ws2)   # 1ª Proposta (com preços)

all_codes = sorted(set(ref.keys()) | set(prop.keys()), key=sort_key)

print(f'Itens no gabarito (Planilha Civil): {len(ref)}')
print(f'Itens na proposta (1ª Proposta):    {len(prop)}')
print()

# ─── Itens só no gabarito ───
only_ref = [c for c in all_codes if c in ref and c not in prop]
if only_ref:
    print(f'=== ITENS SÓ NA PLANILHA CIVIL (ausentes na proposta) [{len(only_ref)}] ===')
    for c in only_ref:
        desc = ref[c]['desc'][:65]
        print(f'  {c:<8} | QDE={ref[c]["qde"]:>6} {ref[c]["unit"]:<5} | {desc}')
    print()

# ─── Itens só na proposta ───
only_prop = [c for c in all_codes if c in prop and c not in ref]
if only_prop:
    print(f'=== ITENS SÓ NA PROPOSTA (novos) [{len(only_prop)}] ===')
    for c in only_prop:
        p = prop[c]
        desc = p['desc'][:55]
        print(f'  {c:<8} | QDE={p["qde"]:>6} {p["unit"]:<5} | MAT={p["mat"]:>10.2f} | MO={p["mo"]:>10.2f} | TOTAL={p["total"]:>12.2f} | {desc}')
    print()

# ─── Diferenças de quantidade ───
print('=== DIFERENÇAS DE QUANTIDADE (Planilha Civil vs 1ª Proposta) ===')
print(f'  {"COD":<8} {"UN":<6} {"QDE_REF":>8} {"QDE_PROP":>9} {"DELTA":>8} | DESCRIÇÃO')
print('  ' + '-' * 95)
diffs = []
for c in all_codes:
    if c not in ref or c not in prop:
        continue
    r, p = ref[c], prop[c]
    qr = r['qde'] if r['qde'] is not None else 0
    qp = p['qde'] if p['qde'] is not None else 0
    if qr != qp:
        delta = qp - qr
        diffs.append((c, r, p, qr, qp, delta))

for c, r, p, qr, qp, delta in diffs:
    flag = ''
    if qr == 0 and qp > 0:
        flag = ' (zerado no ref)'
    elif qp == 0 and qr > 0:
        flag = ' ** ZERADO NA PROP **'
    desc = r['desc'][:45]
    print(f'  {c:<8} {r["unit"]:<6} {qr:>8.2f} {qp:>9.2f} {delta:>+8.2f} | {desc}{flag}')
print(f'\n  Total com diferença de QDE: {len(diffs)}')
print()

# ─── Itens com total = 0 na proposta mas com QDE no gabarito ───
print('=== ITENS SEM PREÇO NA PROPOSTA (total=0 mas QDE ref > 0) ===')
zero_total = []
for c in all_codes:
    if c not in prop:
        continue
    p = prop[c]
    qr = ref.get(c, {}).get('qde') or 0
    if (p['total'] or 0) == 0 and qr > 0:
        zero_total.append((c, p, qr))

for c, p, qr in zero_total:
    desc = p['desc'][:65]
    print(f'  {c:<8} {p["unit"]:<5} QDE_ref={qr:>6} | {desc}')
print(f'\n  Total sem preço: {len(zero_total)}')
print()

# ─── Resumo das seções (tabela resumo, linhas 7–28) ───
print('=== TOTAIS POR SEÇÃO ===')
ref_sec  = {}
prop_sec = {}

for row in ws1.iter_rows(min_row=7, max_row=28, values_only=True):
    desc = str(row[3] or '').strip()
    val  = row[8]
    if desc and val is not None and desc not in ('DESCRIÇÃO', 'TOTAL', ''):
        ref_sec[desc] = val

for row in ws2.iter_rows(min_row=7, max_row=28, values_only=True):
    desc = str(row[3] or '').strip()
    val  = row[8]
    if desc and val is not None and desc not in ('DESCRIÇÃO', 'TOTAL', ''):
        prop_sec[desc] = val

all_sec = sorted(set(ref_sec) | set(prop_sec))
print(f'  {"SEÇÃO":<50} {"GABARITO":>14} {"PROPOSTA":>14}  {"DELTA":>14}')
print('  ' + '-' * 100)
for desc in all_sec:
    rv = ref_sec.get(desc, 0) or 0
    pv = prop_sec.get(desc, 0) or 0
    delta = pv - rv
    flag = '  ** ZERADO **' if pv == 0 and rv != 0 else ''
    print(f'  {desc:<50} {rv:>14,.2f} {pv:>14,.2f}  {delta:>+14,.2f}{flag}')

ref_grand  = sum(v or 0 for v in ref_sec.values())
prop_grand = sum(v or 0 for v in prop_sec.values())
print('  ' + '=' * 100)
print(f'  {"TOTAL GERAL":<50} {ref_grand:>14,.2f} {prop_grand:>14,.2f}  {prop_grand - ref_grand:>+14,.2f}')
