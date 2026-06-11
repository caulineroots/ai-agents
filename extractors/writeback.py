# -*- coding: utf-8 -*-
"""
writeback.py — Fase 5: grava os valores de volta NA PRÓPRIA planilha do cliente,
preservando formatação/fórmulas, e gera o relatório de auditoria por linha.

Edita células (não regenera o workbook) via row_ref capturado no parsing. Itens
'encontrado' recebem QDE; todos os priceáveis recebem TOTAL (v1: combinado em
TOTAL, MAT/M.OBRA em branco). 'divergente' MANTÉM a QDE original — a divergência
vive no relatório, não na célula. Ver docs/arquitetura/06-escrita-e-revisao.md.
"""

import logging

log = logging.getLogger("extractor")


def escrever_planilha(planilha_path: str, itens: list[dict], out_path: str,
                      col_map: dict) -> dict:
    """Preenche uma CÓPIA da planilha. col_map: papel->índice 0-based (do parser)."""
    import openpyxl
    wb = openpyxl.load_workbook(planilha_path)   # mantém formatação (sem data_only)
    ws = wb.active

    col_qde = col_map.get("qde")
    col_tot = col_map.get("total")
    n_qde = n_tot = 0
    for it in itens:
        row = it.get("row_ref")
        if not row:
            continue
        # QDE: só para itens 'encontrado' (planilha não tinha medida)
        if it.get("status") == "encontrado" and it.get("qde_final") is not None and col_qde is not None:
            ws.cell(row=row, column=col_qde + 1, value=round(it["qde_final"], 2))
            n_qde += 1
        # TOTAL: para itens priceados
        preco = it.get("preco")
        if preco and preco.get("total") is not None and col_tot is not None:
            ws.cell(row=row, column=col_tot + 1, value=preco["total"])
            n_tot += 1

    wb.save(out_path)
    log.info("[writeback] %s -> %s | QDE preenchidas=%d, TOTAL=%d", planilha_path, out_path, n_qde, n_tot)
    return {"out_path": out_path, "n_qde_preenchidas": n_qde, "n_total_preenchidos": n_tot}


def relatorio_auditoria(itens: list[dict]) -> list[dict]:
    """Relatório por linha: medida, fonte, método, confiança, status, flags, preço."""
    out = []
    for it in itens:
        m = it.get("medida") or {}
        p = it.get("preco") or {}
        out.append({
            "item": it.get("item"),
            "descricao": it.get("descricao", "")[:80],
            "categoria": it.get("categoria"),
            "estrategia": it.get("estrategia"),
            "status": it.get("status"),
            "qde_final": it.get("qde_final"),
            "unidade": it.get("unidade"),
            "medida_valor": m.get("valor"),
            "metodo": m.get("metodo"),
            "fonte": m.get("fonte"),
            "confianca": m.get("confianca"),
            "preco_total": p.get("total"),
            "preco_match": p.get("match"),
            "preco_ref": p.get("price_ref"),
            "flags": it.get("flags", []),
            "precisa_revisao": bool(set(it.get("flags", [])) & {
                "QTY_MISMATCH", "NOT_FOUND", "LOW_CONFIDENCE", "ESTIMATED",
                "PRECO_INCERTO", "ESTRATEGIA_INDEFINIDA"}),
        })
    return out


def lista_revisao(itens: list[dict]) -> list[dict]:
    """Apenas as linhas que exigem atenção humana (work-list)."""
    return [r for r in relatorio_auditoria(itens) if r["precisa_revisao"]
            or r["status"] in ("divergente", "manual")]
