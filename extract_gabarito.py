"""
Extrai todos os itens de descrição dos Excel de proposta da Celmar
e compara com o banco de nomenclaturas atual.
Gera um relatório de cobertura e itens faltantes.
"""
import os
import re
import json
import unicodedata
import openpyxl
from pathlib import Path
from collections import defaultdict

FOLDER = Path(r"C:\Users\AVELL\Downloads\Cauline Roots\Celmar")
DB_PATH = Path(__file__).parent / "nomenclaturas_db.json"


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s).lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()


def is_junk(v) -> bool:
    if not v or not isinstance(v, str):
        return True
    s = v.strip()
    if not s or len(s) < 3:
        return True
    # Headers e metadados
    skip = {"descricao", "un", "und", "unidade", "item", "codigo", "codigo", "total",
            "totais", "mat.", "m.obra", "mat.*", "m.obra", "c.c.", "loja", "razao social",
            "cnpj", "gerenciador", "duracao da obra", "tipo", "tamanho", "area",
            "previsao inicio da obra", "fim da obra", "omissos", "dias",
            "custos indiretos", "sub-total", "subtotal", "shopping", "celmar"}
    n = norm(s)
    if n in skip:
        return True
    # Só números / códigos
    if re.fullmatch(r"[\d\s\.\-\/\#\*]+", s):
        return True
    # Linhas de separação
    if re.fullmatch(r"[\-\=\_\*\s]+", s):
        return True
    # Muito longo (frases)
    if len(s) > 60:
        return True
    return False


def match_item(name: str, db_items: list) -> dict | None:
    n = norm(name)
    for it in db_items:
        if n == norm(it.get("canonical_name", "")):
            return it
        if any(n in norm(v) or norm(v) in n for v in it.get("variations", [])):
            return it
    return None


def read_descriptions(fpath: Path) -> list[str]:
    """Lê todas as strings de descrição únicas do arquivo Excel."""
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    descs = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and not is_junk(cell):
                    descs.add(cell.strip())
    return list(descs)


def main():
    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    db_items = db["items"]

    all_descs: dict[str, set] = defaultdict(set)  # norm -> {original forms}

    files = [f for f in os.listdir(FOLDER) if f.endswith(".xlsx")]
    print(f"Arquivos encontrados: {len(files)}\n")

    for fname in sorted(files):
        descs = read_descriptions(FOLDER / fname)
        for d in descs:
            all_descs[norm(d)].add(d)

    print(f"Total de descrições únicas extraídas: {len(all_descs)}\n")

    matched = []
    unmatched = []

    for n, forms in sorted(all_descs.items()):
        best = max(forms, key=len)  # forma mais completa
        m = match_item(best, db_items)
        if m:
            matched.append((best, m["id"]))
        else:
            unmatched.append(best)

    print(f"Cobertura atual: {len(matched)}/{len(all_descs)} = {100*len(matched)//len(all_descs)}%\n")

    print("=== RECONHECIDOS ===")
    for name, mid in sorted(matched, key=lambda x: x[1]):
        print(f"  [{mid}] {name}")

    print(f"\n=== NÃO MAPEADOS ({len(unmatched)}) — candidatos para o banco ===")
    for name in sorted(unmatched):
        print(f"  {name!r}")

    # Salva resultado
    result = {
        "total_unique": len(all_descs),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "coverage_pct": round(100*len(matched)/len(all_descs), 1),
        "unmatched_items": sorted(unmatched),
        "matched_items": [{"name": n, "matched_to": m} for n, m in matched],
    }
    out = Path(__file__).parent / "gabarito_analysis.json"
    out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nResultado salvo em: {out}")


if __name__ == "__main__":
    main()
