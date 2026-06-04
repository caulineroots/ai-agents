import openpyxl, glob, json

files = glob.glob(r'C:\Users\AVELL\Downloads\Cauline Roots\Celmar\*.xlsx')
print("Arquivos encontrados:", [f.split("\\")[-1] for f in files])

wb = openpyxl.load_workbook(files[0], data_only=True)
ws = wb[wb.sheetnames[0]]

print("\n=== SUMÁRIO ===")
for i, row in enumerate(ws.iter_rows(min_row=7, max_row=29, values_only=True)):
    if any(v is not None for v in row):
        print(f"L{i+7}:", row)

print("\n=== ITENS DETALHADOS ===")
items = []
for i, row in enumerate(ws.iter_rows(min_row=30, max_row=ws.max_row, values_only=True)):
    if not any(v is not None for v in row):
        continue
    cols = list(row) + [None]*12
    cc, _, item, desc, un, qde, mat, mo, total = cols[:9]
    if (desc or total) and total and isinstance(total, (int, float)) and total > 0:
        items.append({
            "cc": cc,
            "item": str(item or ""),
            "descricao": str(desc or "")[:80],
            "un": str(un or ""),
            "qde": qde,
            "material": mat,
            "mao_obra": mo,
            "total": total,
        })
        print(f"  [{cc or '':>6}] {str(item or ''):>6}  {str(desc or '')[:60]:<60}  {str(un or ''):>4}  qde={str(qde or ''):>8}  R${total:>12.2f}")

# Totals by category
print("\n=== TOTAL POR GRUPO ===")
grupos = {}
for it in items:
    desc_lower = (it['descricao'] or '').lower()
    if 'prov' in desc_lower or 'espelho' in desc_lower:
        g = 'PROVADORES'
    elif 'forro' in desc_lower:
        g = 'FORRO'
    elif 'piso' in desc_lower:
        g = 'PISO'
    elif 'pintur' in desc_lower:
        g = 'PINTURA'
    elif 'alvenar' in desc_lower or 'drywall' in desc_lower or 'divisor' in desc_lower:
        g = 'CIVIL'
    elif 'painel' in desc_lower or 'ferrag' in desc_lower:
        g = 'PAINEIS'
    elif 'fachad' in desc_lower or 'vitrin' in desc_lower or 'caixilh' in desc_lower:
        g = 'FACHADA'
    elif 'admin' in desc_lower or 'mobili' in desc_lower or 'limpez' in desc_lower or 'vigiil' in desc_lower:
        g = 'ADMIN'
    else:
        g = 'OUTROS'
    grupos[g] = grupos.get(g, 0) + it['total']

for g, v in sorted(grupos.items(), key=lambda x: -x[1]):
    print(f"  {g:<20} R${v:>12.2f}")

total = sum(it['total'] for it in items)
print(f"\n  TOTAL GERAL          R${total:>12.2f}")
